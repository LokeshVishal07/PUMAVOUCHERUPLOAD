import streamlit as st
import pandas as pd
import re
import io
import zipfile
from openpyxl import load_workbook

# ─────────────────────────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="PUMA Voucher SKU Tool",
    page_icon="🏷️",
    layout="wide"
)

# ─────────────────────────────────────────────────────────────────
# CONSTANTS  (no more hardcoded column indices — user selects them)
# ─────────────────────────────────────────────────────────────────

REGION_MARKETPLACES = {
    "PH": ["Lazada", "Shopee", "Zalora"],
    "MY": ["Lazada", "Shopee", "Zalora", "TikTok"],
    "SG": ["Lazada", "Shopee", "Zalora"],
}

REGION_CONFIG = {
    "PH": {
        "zecom_sheet":  "PH",
        "zecom_read":   "ph",
        "article_col":  "PIM Article#",
        "threshold":    650,
        "currency":     "PHP > 650",
        "mp_flags": {
            "Lazada": "LAZADA",
            "Shopee": "SHOPEE",
            "Zalora": "ZALORA",
        },
        # Fallback column indices (used ONLY as initial dropdown default)
        "default_excl": 71,
        "default_rrp":  32,
        "default_srp":  50,
    },
    "MY": {
        "zecom_sheet":  "MY",
        "zecom_read":   "header3",
        "article_col":  "Style#",
        "threshold":    36,
        "currency":     "RM > 36",
        "mp_flags": {
            "Lazada":  "Lazada",
            "Shopee":  "Shopee",
            "Zalora":  "Zalora MP",
            "TikTok":  "TIKTOK",
        },
        # Lazada/Zalora/TikTok block defaults; Shopee differs — user selects
        "default_excl": 51,
        "default_rrp":  48,
        "default_srp":  49,
    },
    "SG": {
        "zecom_sheet":  "SG",
        "zecom_read":   "header3",
        "article_col":  "STYLE#",
        "threshold":    16,
        "currency":     "SGD > 16",
        "mp_flags": {
            "Lazada": "Lazada",
            "Shopee": "Shopee",
            "Zalora": "Zalora",
        },
        "default_excl": 52,
        "default_rrp":  26,
        "default_srp":  50,
    },
}

# ─────────────────────────────────────────────────────────────────
# ZECOM READING & VALIDATION
# ─────────────────────────────────────────────────────────────────

def validate_zecom_region(file_bytes: bytes, selected_region: str):
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception as e:
        return False, f"Cannot read ZeCom file: {e}"

    if selected_region == "PH":
        if "PH" not in sheets:
            extra = " (looks like MY/SG tracker)" if ("MY" in sheets or "SG" in sheets) else ""
            return False, f"⚠️ Wrong file! Selected **PH** but 'PH' sheet not found{extra}."
    else:
        if selected_region not in sheets:
            extra = " (looks like PH tracker)" if "PH" in sheets else ""
            return False, f"⚠️ Wrong file! Selected **{selected_region}** but '{selected_region}' sheet not found{extra}."
    return True, "OK"


@st.cache_data(show_spinner=False)
def read_zecom(file_bytes: bytes, region: str) -> pd.DataFrame:
    cfg = REGION_CONFIG[region]
    sheet = cfg["zecom_sheet"]
    if cfg["zecom_read"] == "ph":
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name="PH", header=1)
        df.columns = df.iloc[0]
        df = df.iloc[1:].reset_index(drop=True)
    else:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet, header=3)
    return df


# ─────────────────────────────────────────────────────────────────
# COLUMN DETECTION HELPERS
# ─────────────────────────────────────────────────────────────────

_EXCL_RE = re.compile(
    r'open for all|exclude|vc only|vc max|vc -|shopee exclusive|platform vc',
    re.I
)
_RRP_WORDS = ["rrp", "retail price"]
_SRP_WORDS = ["srp ao", "ec srp", "srp"]


def col_options(df: pd.DataFrame) -> list[str]:
    """Formatted column labels: '[idx] col_name'."""
    return [f"[{i}]  {col}" for i, col in enumerate(df.columns)]


def _col_score_excl(series: pd.Series) -> int:
    """Count how many non-null values match exclusion remark patterns."""
    vals = series.dropna().astype(str).str.strip()
    vals = vals[vals != ""]
    if vals.empty:
        return 0
    return int(vals.apply(lambda v: bool(_EXCL_RE.search(v))).sum())


def _col_by_name(df: pd.DataFrame, keywords: list[str]) -> int | None:
    """First column whose name contains any keyword (case-insensitive)."""
    for kw in keywords:
        for i, col in enumerate(df.columns):
            if kw.lower() in str(col).lower():
                return i
    return None


def guess_excl_idx(df: pd.DataFrame, fallback: int) -> int:
    """Return the column index most likely to be the exclusion/campaign column."""
    best_col, best_score = fallback, 0
    for i in range(len(df.columns)):
        s = _col_score_excl(df.iloc[:, i])
        if s > best_score:
            best_score, best_col = s, i
    return best_col


def guess_rrp_idx(df: pd.DataFrame, fallback: int) -> int:
    idx = _col_by_name(df, _RRP_WORDS)
    return idx if idx is not None else fallback


def guess_srp_idx(df: pd.DataFrame, fallback: int) -> int:
    # Find by name but skip pure "rrp" hits
    for kw in _SRP_WORDS:
        for i, col in enumerate(df.columns):
            col_str = str(col).lower()
            if kw in col_str and "rrp" not in col_str:
                return i
    return fallback


def sample_vals(df: pd.DataFrame, col_idx: int, n: int = 6) -> str:
    """Return a comma-joined string of sample values from the column."""
    vals = df.iloc[:, col_idx].dropna()
    uniq = vals.unique()[:n]
    return ", ".join(str(v) for v in uniq) if len(uniq) else "(no values)"


# ─────────────────────────────────────────────────────────────────
# INVENTORY READING
# ─────────────────────────────────────────────────────────────────

def _normalize_inv(df, ean_cands, stock_cands):
    ean_col   = next((c for c in ean_cands   if c in df.columns), None)
    stock_col = next((c for c in stock_cands if c in df.columns), None)
    if ean_col is None or stock_col is None:
        return None
    out = df[[ean_col, stock_col]].copy()
    out.columns = ["EAN", "Stock"]
    out["EAN"]   = out["EAN"].astype(str).str.strip()
    out["Stock"] = pd.to_numeric(out["Stock"], errors="coerce").fillna(0)
    out = out[out["EAN"].str.match(r"^\d{13}$")]
    return out


def read_inventory(file_bytes: bytes, filename: str):
    ean_c   = ["EAN", "Sku", "PROD_CODE", "SellerSku"]
    stock_c = ["Avail_Qty", "QtyAvailable", "QTY", "Quantity"]
    if filename.lower().endswith(".csv"):
        return _normalize_inv(pd.read_csv(io.BytesIO(file_bytes)), ean_c, stock_c)
    r = _normalize_inv(pd.read_excel(io.BytesIO(file_bytes)), ean_c, stock_c)
    if r is not None:
        return r
    return _normalize_inv(pd.read_excel(io.BytesIO(file_bytes), header=4), ean_c, stock_c)


# ─────────────────────────────────────────────────────────────────
# VOUCHER ELIGIBILITY LOGIC
# ─────────────────────────────────────────────────────────────────

def classify_remark(remark, voucher_pct: int, voucher_type: str,
                    marketplace: str = None) -> str:
    """
    Returns: 'eligible' | 'ineligible' | 'no_remark'

    Bundle discount: only OPEN FOR ALL* remarks qualify.
    Regular VC:
      OPEN FOR ALL*              → any %
      OPEN for N days up to Y%  → any % ≤ Y
      N% VC ONLY / MAX / - *    → exactly N%
      N% VC - Shopee exclusive   → Shopee only at exactly N%
      EXCLUDE*                   → never
      blank / null               → no_remark
    """
    if pd.isna(remark):
        return "no_remark"
    r = str(remark).strip()
    if not r or r.lower() == "nan":
        return "no_remark"
    rl = r.lower()

    if rl.startswith("exclude"):
        return "ineligible"

    if rl.startswith("open for all"):          # covers all OPEN FOR ALL variants
        return "eligible"

    if voucher_type == "bundle":               # bundle: only OPEN FOR ALL qualifies
        return "ineligible"

    # "OPEN for N days up to Y%"
    m = re.search(r"open for.*?up to\s*(\d+)%", rl)
    if m:
        return "eligible" if voucher_pct <= int(m.group(1)) else "ineligible"

    # Exact % match in remark
    m = re.search(r"(\d+)%", r)
    if m:
        remark_pct = int(m.group(1))
        if remark_pct == voucher_pct:
            if "shopee exclusive" in rl:
                return "eligible" if (marketplace and "shopee" in marketplace.lower()) \
                       else "ineligible"
            return "eligible"
        return "ineligible"

    return "ineligible"


# ─────────────────────────────────────────────────────────────────
# ZECOM ARTICLE-LEVEL PROCESSING   (column indices passed explicitly)
# ─────────────────────────────────────────────────────────────────

def process_zecom(zecom_df, region, marketplace,
                  excl_idx, rrp_idx, srp_idx,
                  voucher_pct, voucher_type) -> pd.DataFrame:
    """
    Filter ZeCom by MP flag, price threshold, and remark.
    Returns DataFrame [article, remark_status].
    """
    cfg = REGION_CONFIG[region]
    df  = zecom_df.copy()

    # MP flag filter
    mp_col = cfg["mp_flags"].get(marketplace)
    if mp_col and mp_col in df.columns:
        df = df[df[mp_col].astype(str).str.strip().str.upper() == "YES"].copy()

    if df.empty:
        return pd.DataFrame(columns=["article", "remark_status"])

    threshold = cfg["threshold"]
    rrp = pd.to_numeric(df.iloc[:, rrp_idx], errors="coerce")
    srp = pd.to_numeric(df.iloc[:, srp_idx], errors="coerce")
    price_ok = (rrp > threshold) & (srp > threshold)

    excl_vals = df.iloc[:, excl_idx]

    result = pd.DataFrame({
        "article":  df[cfg["article_col"]].astype(str).str.strip().values,
        "price_ok": price_ok.values,
        "remark":   excl_vals.values,
    })

    result["remark_status"] = result.apply(
        lambda row: "ineligible" if not row["price_ok"]
        else classify_remark(row["remark"], voucher_pct, voucher_type, marketplace),
        axis=1,
    )

    result = result[result["article"].str.match(r"^[\w_\-]+$", na=False)]
    result = result[result["article"].str.lower() != "nan"]
    return result[["article", "remark_status"]].drop_duplicates(subset=["article"])


# ─────────────────────────────────────────────────────────────────
# EAN MAPPING & ELIGIBILITY SETS
# ─────────────────────────────────────────────────────────────────

def map_to_eans(article_status, content_df, inventory_df):
    merged = article_status.merge(
        content_df.rename(columns={"Color_No": "article"}),
        on="article", how="inner",
    )
    merged["EAN"] = merged["EAN"].astype(str).str.strip()
    merged = merged.merge(inventory_df.rename(columns={"Stock": "stock"}),
                          on="EAN", how="left")
    merged["has_stock"] = merged["stock"].fillna(0) > 0
    return merged[["article", "EAN", "remark_status", "has_stock"]]


def eligible_ean_set(ean_df):
    return set(ean_df[(ean_df["remark_status"] == "eligible") & ean_df["has_stock"]]["EAN"])


def excluded_ean_set(ean_df):
    return set(ean_df[ean_df["remark_status"] == "ineligible"]["EAN"])


def no_remark_ean_set(ean_df):
    return set(ean_df[ean_df["remark_status"] == "no_remark"]["EAN"])


# ─────────────────────────────────────────────────────────────────
# MARKETPLACE PROCESSORS
# ─────────────────────────────────────────────────────────────────

def process_lazada(ean_df, lazada_bytes):
    df   = pd.read_excel(io.BytesIO(lazada_bytes), sheet_name="template", header=0)
    data = df.iloc[3:].copy()
    data.columns = df.columns
    active = data[data["status"].astype(str).str.lower() == "active"].copy()
    active["_ean"] = active["SellerSKU"].astype(str).str.strip()
    ok = eligible_ean_set(ean_df)
    return active[active["_ean"].isin(ok)]["Shop SKU"].dropna().unique().tolist()


def _read_shopee_zip(zip_bytes):
    dfs = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        names = sorted(n for n in zf.namelist() if n.endswith(".xlsx"))
        bar = st.progress(0, text="Reading Shopee export files…")
        for i, name in enumerate(names):
            with zf.open(name) as f:
                dfs.append(pd.read_excel(f, engine="calamine", header=2, skiprows=[3, 4]))
            bar.progress((i + 1) / len(names), text=f"Reading Shopee file {i+1}/{len(names)}…")
        bar.empty()
    return pd.concat(dfs, ignore_index=True)


def _extract_ean(sku_val, parent_val):
    for v in (sku_val, parent_val):
        if pd.notna(v):
            try:
                s = str(int(float(v)))
                if re.match(r"^\d{13}$", s):
                    return s
            except (ValueError, TypeError):
                s = str(v).strip()
                if re.match(r"^\d{13}$", s):
                    return s
    return None


def _pid_eligible(combined, ok_eans, excl_eans):
    combined = combined.copy()
    combined["_ean"] = combined.apply(
        lambda r: _extract_ean(r.get("SKU"), r.get("Parent SKU")), axis=1
    )
    combined["_pid"] = combined["Product ID"].astype(str).str.strip()
    result = []
    for pid, grp in combined.groupby("_pid"):
        g = set(grp["_ean"].dropna())
        if g & excl_eans:
            continue
        if g & ok_eans:
            result.append(pid)
    return result


def process_shopee(ean_df, zip_bytes):
    combined = _read_shopee_zip(zip_bytes)
    return _pid_eligible(combined, eligible_ean_set(ean_df), excluded_ean_set(ean_df))


def process_zalora(ean_df, eligible_bytes, content_df):
    df = pd.read_excel(io.BytesIO(eligible_bytes), sheet_name="Eligible Products")
    ok  = eligible_ean_set(ean_df)
    nr  = no_remark_ean_set(ean_df)
    ean2art = dict(zip(content_df["EAN"].astype(str).str.strip(),
                       content_df["Color_No"].astype(str).str.strip()))
    df["_ean"]       = df["Seller SKU"].astype(str).str.strip()
    df["Article No"] = df["_ean"].map(ean2art)
    df["Voucher Eligible"] = df["_ean"].apply(
        lambda e: "Yes" if e in ok else ("No Remark" if e in nr else "No")
    )
    df.drop(columns=["_ean"], inplace=True)
    return df


def process_tiktok(ean_df, tiktok_bytes):
    df = pd.read_excel(io.BytesIO(tiktok_bytes), sheet_name="Template",
                       header=2, skiprows=[3, 4])
    df["_ean"] = df["Seller SKU"].apply(lambda v: _extract_ean(v, None))
    df["_pid"] = df["Product ID"].astype(str).str.strip()
    ok  = eligible_ean_set(ean_df)
    exc = excluded_ean_set(ean_df)
    result = []
    for pid, grp in df.groupby("_pid"):
        g = set(grp["_ean"].dropna())
        if g & exc:
            continue
        if g & ok:
            result.append(pid)
    return result


# ─────────────────────────────────────────────────────────────────
# OUTPUT HELPERS
# ─────────────────────────────────────────────────────────────────

def _to_excel(df, sheet="Sheet1"):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet)
    return buf.getvalue()


def make_lazada_output(ids):  return _to_excel(pd.DataFrame({"SHOP SKU": ids}))
def make_shopee_output(ids):  return _to_excel(pd.DataFrame({"Product ID": ids}))
def make_zalora_output(ann):  return _to_excel(ann, sheet="Eligible Products")


# ─────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────

def main():
    st.title("🏷️ PUMA Voucher Eligible SKU Tool")
    st.caption("Generate marketplace-ready voucher eligible SKU lists from ZeCom tracker data.")

    # ── ① REGION & MARKETPLACE ───────────────────────────────────
    st.markdown("---")
    st.subheader("① Region & Marketplace")
    c1, c2 = st.columns(2)
    with c1:
        region = st.selectbox("Region", ["PH", "MY", "SG"])
    with c2:
        marketplace = st.selectbox("Marketplace", REGION_MARKETPLACES[region])

    # ── ② VOUCHER CONFIG ─────────────────────────────────────────
    st.markdown("---")
    st.subheader("② Voucher Configuration")
    cv1, cv2 = st.columns([1, 2])
    with cv1:
        voucher_type = st.radio("Voucher Type", ["Regular VC", "Bundle Discount"], horizontal=True)
    with cv2:
        pct_raw = st.text_input("Voucher % — comma-separated (e.g. 10, 20, 50)",
                                 placeholder="10, 20, 50")

    voucher_pcts = [int(t) for t in re.split(r"[,\s%]+", pct_raw.strip()) if t.isdigit()]
    if voucher_pcts:
        vt = "Bundle Discount" if voucher_type == "Bundle Discount" else "Regular VC"
        st.success(f"Will generate: **{', '.join(str(p)+'%' for p in voucher_pcts)} {vt}** "
                   f"— **{region} / {marketplace}**")

    # ── ③ FILE UPLOADS ───────────────────────────────────────────
    st.markdown("---")
    st.subheader("③ Upload Files")

    cf1, cf2 = st.columns(2)
    with cf1:
        st.markdown("**Core files (required for all)**")
        zecom_file   = st.file_uploader(
            f"ZeCom Tracker ({'PH file' if region == 'PH' else 'MY + SG file'})",
            type=["xlsx"], key="zecom"
        )
        content_file = st.file_uploader("Content File", type=["xlsx"], key="content")
        inv_file     = st.file_uploader(f"Inventory File ({region})",
                                         type=["xlsx", "csv"], key="inv")

    with cf2:
        st.markdown(f"**{marketplace} marketplace export**")
        if marketplace == "Lazada":
            mp_file = st.file_uploader("Lazada Product Export (.xlsx)",
                                        type=["xlsx"], key="mp")
        elif marketplace == "Shopee":
            mp_file = st.file_uploader("Shopee Export ZIP (.zip — all batches)",
                                        type=["zip"], key="mp")
        elif marketplace == "Zalora":
            mp_file = st.file_uploader("Zalora EligibleProducts File (.xlsx)",
                                        type=["xlsx"], key="mp")
        elif marketplace == "TikTok":
            mp_file = st.file_uploader("TikTok Seller Center Export (.xlsx)",
                                        type=["xlsx"], key="mp")
        else:
            mp_file = None

    # ── DYNAMIC COLUMN SELECTORS (appear after ZeCom upload) ─────
    excl_idx = rrp_idx = srp_idx = None

    if zecom_file:
        ok_region, err_msg = validate_zecom_region(zecom_file.getvalue(), region)
        if not ok_region:
            st.error(err_msg)
            st.stop()

        st.markdown("---")
        st.subheader("③b  Select ZeCom Columns")
        st.caption(
            "The app has pre-selected the most likely columns. "
            "Check the **sample values** and change the dropdown if needed — "
            "especially useful when the campaign block has shifted."
        )

        with st.spinner("Reading ZeCom file…"):
            zecom_df = read_zecom(zecom_file.getvalue(), region)

        cfg      = REGION_CONFIG[region]
        opts     = col_options(zecom_df)
        n_cols   = len(zecom_df.columns)

        def safe_idx(i): return min(i, n_cols - 1)

        # Smart defaults
        default_excl = safe_idx(guess_excl_idx(zecom_df, cfg["default_excl"]))
        default_rrp  = safe_idx(guess_rrp_idx (zecom_df, cfg["default_rrp"]))
        default_srp  = safe_idx(guess_srp_idx (zecom_df, cfg["default_srp"]))

        sc1, sc2, sc3 = st.columns(3)

        with sc1:
            st.markdown("**📋 Exclusion / Campaign Column**")
            excl_sel = st.selectbox(
                "Select exclusion column",
                opts,
                index=default_excl,
                key="sel_excl",
                label_visibility="collapsed",
            )
            excl_idx = int(re.search(r"\[(\d+)\]", excl_sel).group(1))
            sv = sample_vals(zecom_df, excl_idx)
            st.caption(f"Sample values: `{sv}`")

        with sc2:
            st.markdown("**💰 RRP Column**")
            rrp_sel = st.selectbox(
                "Select RRP column",
                opts,
                index=default_rrp,
                key="sel_rrp",
                label_visibility="collapsed",
            )
            rrp_idx = int(re.search(r"\[(\d+)\]", rrp_sel).group(1))
            sv = sample_vals(zecom_df, rrp_idx)
            st.caption(f"Sample values: `{sv}`")

        with sc3:
            st.markdown("**🏷️ SRP Column**")
            srp_sel = st.selectbox(
                "Select SRP column",
                opts,
                index=default_srp,
                key="sel_srp",
                label_visibility="collapsed",
            )
            srp_idx = int(re.search(r"\[(\d+)\]", srp_sel).group(1))
            sv = sample_vals(zecom_df, srp_idx)
            st.caption(f"Sample values: `{sv}`")

        # Price threshold reminder
        st.info(
            f"Price filter: both **RRP** and **SRP** must be **> {cfg['threshold']} "
            f"({cfg['currency']})** to be eligible."
        )

    # ── ④ GENERATE ───────────────────────────────────────────────
    st.markdown("---")
    st.subheader("④ Generate")

    missing = []
    if not zecom_file:   missing.append("ZeCom Tracker")
    if not content_file: missing.append("Content File")
    if not inv_file:     missing.append("Inventory File")
    if not mp_file:      missing.append(f"{marketplace} Export")
    if not voucher_pcts: missing.append("Voucher %")
    if excl_idx is None: missing.append("Column selection (upload ZeCom first)")

    if missing:
        st.info(f"Still needed: **{', '.join(missing)}**")

    ready = not missing

    if st.button("🚀 Generate Eligible SKU Lists", disabled=not ready, type="primary"):
        _run(
            zecom_file, content_file, inv_file, mp_file,
            region, marketplace,
            excl_idx, rrp_idx, srp_idx,
            voucher_pcts, voucher_type,
        )


# ─────────────────────────────────────────────────────────────────
# PROCESSING PIPELINE
# ─────────────────────────────────────────────────────────────────

def _run(zecom_file, content_file, inv_file, mp_file,
         region, marketplace,
         excl_idx, rrp_idx, srp_idx,
         voucher_pcts, voucher_type):

    vtype = "bundle" if voucher_type == "Bundle Discount" else "regular"

    with st.status("Processing…", expanded=True) as status:

        # 1. Read ZeCom (already cached from column selector above)
        st.write(f"📊 Reading ZeCom — {region}…")
        try:
            zecom_df = read_zecom(zecom_file.getvalue(), region)
            st.write(f"   ✓ {len(zecom_df):,} rows | "
                     f"excl=col[{excl_idx}]  rrp=col[{rrp_idx}]  srp=col[{srp_idx}]")
        except Exception as e:
            st.error(f"ZeCom read error: {e}")
            status.update(label="❌ Error", state="error"); return

        # 2. Read Content
        st.write("📦 Reading content file…")
        try:
            content_df = pd.read_excel(io.BytesIO(content_file.getvalue()), sheet_name="content")
            content_df = content_df[["Color_No", "EAN"]].dropna()
            content_df["EAN"]      = content_df["EAN"].astype(str).str.strip()
            content_df["Color_No"] = content_df["Color_No"].astype(str).str.strip()
            st.write(f"   ✓ {len(content_df):,} EAN mappings")
        except Exception as e:
            st.error(f"Content file error: {e}")
            status.update(label="❌ Error", state="error"); return

        # 3. Read Inventory
        st.write("🏭 Reading inventory…")
        try:
            inv_df = read_inventory(inv_file.getvalue(), inv_file.name)
            if inv_df is None:
                st.error("Could not detect EAN/Stock columns. "
                         "Expected: EAN/Sku/PROD_CODE  and  Avail_Qty/QtyAvailable/QTY.")
                status.update(label="❌ Error", state="error"); return
            st.write(f"   ✓ {len(inv_df):,} EANs | {(inv_df['Stock']>0).sum():,} in stock")
        except Exception as e:
            st.error(f"Inventory read error: {e}")
            status.update(label="❌ Error", state="error"); return

        # 4. Process each voucher %
        results = {}
        for pct in voucher_pcts:
            st.write(f"⚙️  Processing **{pct}% {voucher_type}**…")

            art = process_zecom(zecom_df, region, marketplace,
                                excl_idx, rrp_idx, srp_idx, pct, vtype)
            n_elig = (art["remark_status"] == "eligible").sum()
            n_nr   = (art["remark_status"] == "no_remark").sum()
            n_ineli= (art["remark_status"] == "ineligible").sum()
            st.write(f"   ZeCom: {n_elig} eligible | {n_nr} no-remark | {n_ineli} ineligible")

            if n_elig + n_nr == 0:
                st.warning(f"   No eligible / no-remark articles for {pct}% — skipping.")
                continue

            ean_df = map_to_eans(art, content_df, inv_df)
            n_ok   = len(eligible_ean_set(ean_df))
            st.write(f"   EANs in stock & eligible: {n_ok:,}")

            if n_ok == 0:
                st.warning(f"   No in-stock eligible EANs for {pct}% — skipping.")
                continue

            try:
                if marketplace == "Lazada":
                    ids = process_lazada(ean_df, mp_file.getvalue())
                    st.write(f"   ✅ Lazada Shop SKUs → **{len(ids)}**")
                    results[pct] = {"mp": "Lazada", "ids": ids}

                elif marketplace == "Shopee":
                    ids = process_shopee(ean_df, mp_file.getvalue())
                    st.write(f"   ✅ Shopee Product IDs → **{len(ids)}**")
                    results[pct] = {"mp": "Shopee", "ids": ids}

                elif marketplace == "Zalora":
                    ann = process_zalora(ean_df, mp_file.getvalue(), content_df)
                    y  = (ann["Voucher Eligible"] == "Yes").sum()
                    nr = (ann["Voucher Eligible"] == "No Remark").sum()
                    st.write(f"   ✅ Zalora: {y} eligible | {nr} no-remark")
                    results[pct] = {"mp": "Zalora", "ann": ann, "yes_count": y}

                elif marketplace == "TikTok":
                    ids = process_tiktok(ean_df, mp_file.getvalue())
                    st.write(f"   ✅ TikTok Product IDs → **{len(ids)}**")
                    results[pct] = {"mp": "TikTok", "ids": ids}

            except Exception as e:
                st.error(f"   Error processing {marketplace} for {pct}%: {e}")

        if results:
            status.update(label="✅ Done!", state="complete")
        else:
            status.update(label="⚠️ No results generated", state="error"); return

    # ── ⑤ DOWNLOAD RESULTS ───────────────────────────────────────
    st.markdown("---")
    st.subheader("⑤ Download Results")
    today    = pd.Timestamp.now().strftime("%Y%m%d")
    vt_short = "Bundle" if vtype == "bundle" else "VC"

    for pct, res in results.items():
        mp    = res["mp"]
        fname = f"{mp}_{region}_{pct}pct_{vt_short}_{today}.xlsx"
        cm, cd = st.columns([3, 2])

        with cm:
            if mp == "Zalora":
                st.metric(f"{mp} — {pct}% {vt_short}", f"{res['yes_count']} eligible SKUs")
            else:
                label = "Shop SKUs" if mp == "Lazada" else "Product IDs"
                st.metric(f"{mp} — {pct}% {vt_short}", f"{len(res['ids'])} {label}")

        with cd:
            if mp == "Lazada":      data = make_lazada_output(res["ids"])
            elif mp in ("Shopee", "TikTok"): data = make_shopee_output(res["ids"])
            else:                   data = make_zalora_output(res["ann"])

            st.download_button(
                f"⬇️ Download {fname}", data=data, file_name=fname,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"dl_{mp}_{pct}",
            )


if __name__ == "__main__":
    main()
